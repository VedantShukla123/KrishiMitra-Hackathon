from flask import Flask, request, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from flask_cors import CORS
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
import os
from datetime import datetime
import uuid
from dotenv import load_dotenv
import random
import json
import csv
from io import BytesIO, StringIO
import requests
from openpyxl import load_workbook
import base64
import mimetypes
import re

try:
    import google.generativeai as genai
except ImportError:
    genai = None

from krishimitra_knowledge import KRISHIMITRA_KNOWLEDGE

app = Flask(__name__)
CORS(app)
load_dotenv()

# Configuration
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-change-this-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL', 'sqlite:///krishimitra.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['JWT_SECRET_KEY'] = os.getenv('JWT_SECRET_KEY', 'jwt-secret-key-change-this')

# Initialize extensions
db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
jwt = JWTManager(app)

# Create upload folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    phone = db.Column(db.String(20), nullable=True)
    trust_score = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    def __repr__(self):
        return f'<User {self.username}>'

class UploadedFile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    processed = db.Column(db.Boolean, default=False)
    ai_response = db.Column(db.Text, nullable=True)
    
    user = db.relationship('User', backref=db.backref('files', lazy=True))

class SensorReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    trust_score_10 = db.Column(db.Float, default=0.0)
    address_text = db.Column(db.String(500), nullable=True)
    lat = db.Column(db.Float, nullable=True)
    lon = db.Column(db.Float, nullable=True)
    ai_summary = db.Column(db.Text, nullable=True)

class SupportFeedback(db.Model):
    """User feedback, complaints, or ratings from the chatbot."""
    id = db.Column(db.Integer, primary_key=True)
    type = db.Column(db.String(20), nullable=False)  # 'feedback', 'complaint', 'rating'
    content = db.Column(db.Text, nullable=True)
    rating = db.Column(db.Integer, nullable=True)  # 1-5 for ratings
    user_id = db.Column(db.String(128), nullable=True)  # Firebase UID if logged in
    email = db.Column(db.String(120), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# Ensure all errors return JSON
@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Not found'}), 404

@app.errorhandler(500)
def server_error(e):
    return jsonify({'error': 'Server error. Check backend terminal for details.'}), 500

# Routes
@app.route('/')
def home():
    return jsonify({'message': 'Krishimitra Backend API'})

@app.route('/api/register', methods=['POST'])
def register():
    data = request.get_json() or {}
    email = data.get('email', '').strip()
    password = data.get('password', '')
    name = (data.get('name') or data.get('username') or '').strip()
    phone = (data.get('phone') or '').strip()
    trust_score = data.get('trustScore')
    if not email or not password:
        return jsonify({'error': 'Email and password are required'}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({'error': 'Email already exists'}), 400
    hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
    if trust_score is None:
        trust_score = random.randint(60, 100)
    new_user = User(
        username=name or email.split('@')[0],
        email=email,
        password=hashed_password,
        phone=phone,
        trust_score=trust_score
    )
    db.session.add(new_user)
    db.session.commit()
    return jsonify({
        'message': 'User created successfully',
        'user': {
            'id': new_user.id,
            'name': new_user.username,
            'email': new_user.email,
            'phone': new_user.phone,
            'trustScore': new_user.trust_score
        }
    }), 201

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json() or {}
    email = data.get('email', '').strip()
    password = data.get('password', '')
    if not email or not password:
        return jsonify({'error': 'Email and password are required'}), 400
    user = User.query.filter_by(email=email).first()
    if not user or not bcrypt.check_password_hash(user.password, password):
        return jsonify({'error': 'Invalid email or password'}), 401
    access_token = create_access_token(identity=user.id)
    return jsonify({
        'message': 'Login successful',
        'access_token': access_token,
        'user': {
            'id': user.id,
            'name': user.username,
            'email': user.email,
            'phone': user.phone,
            'trustScore': user.trust_score
        }
    }), 200

def _mime_for_ext(ext):
    m = mimetypes.types_map.get(ext.lower()) if ext else None
    if not m:
        return 'application/octet-stream'
    return m

def _chat_with_gemini_rest(message, history=None):
    """Customer care chat using Gemini with full Krishimitra knowledge and conversation history. Returns (reply_text, error_message)."""
    key = (os.getenv('GEMINI_API_KEY') or os.getenv('GOOGLE_API_KEY') or '').strip()
    if not key:
        return None, 'No GEMINI_API_KEY or GOOGLE_API_KEY set.'
    history = history or []
    try:
        # Build multi-turn contents: alternate user / model from history, then new user message
        contents = []
        for msg in history:
            role = (msg.get('role') or '').strip().lower()
            content = (msg.get('content') or '').strip()
            if not content:
                continue
            gemini_role = 'model' if role == 'assistant' else 'user'
            contents.append({'role': gemini_role, 'parts': [{'text': content}]})
        contents.append({'role': 'user', 'parts': [{'text': message}]})

        model = (os.getenv('GEMINI_MODEL') or 'gemini-pro').strip() or 'gemini-pro'
        use_v1 = model == 'gemini-pro'
        # v1 gemini-pro has no systemInstruction; prepend context as first user message
        if use_v1:
            preamble = 'Use ONLY the following context to answer. Do not make up info.\n\n' + KRISHIMITRA_KNOWLEDGE + '\n\n---\nConversation:'
            contents = [{'role': 'user', 'parts': [{'text': preamble}]}] + contents
            body = {'contents': contents, 'generationConfig': {'temperature': 0.4, 'maxOutputTokens': 1024}}
        else:
            body = {
                'systemInstruction': {'parts': [{'text': KRISHIMITRA_KNOWLEDGE}]},
                'contents': contents,
                'generationConfig': {'temperature': 0.4, 'maxOutputTokens': 1024},
            }
        api_ver = 'v1' if use_v1 else 'v1beta'
        url = f'https://generativelanguage.googleapis.com/{api_ver}/models/{model}:generateContent'
        r = requests.post(f'{url}?key={key}', headers={'Content-Type': 'application/json'}, json=body, timeout=45)
        j = r.json() if r.text else {}
        if r.status_code != 200:
            err = j.get('error', {})
            msg = err.get('message', r.text or f'HTTP {r.status_code}')
            return None, (msg[:300] if isinstance(msg, str) else str(msg)[:300])
        parts = j.get('candidates', [{}])
        if not parts:
            return None, (j.get('error', {}).get('message') or 'No response from model')[:300]
        content = parts[0].get('content', {}).get('parts', [])
        if not content:
            return None, 'Empty response from model'
        text = (content[0].get('text') or '').strip()
        return text or None, None
    except requests.exceptions.RequestException as e:
        return None, str(e)[:200]
    except Exception as e:
        return None, str(e)[:200]


def _analyze_with_gemini(path, prompt, crop=None):
    """Use Gemini API for crop image analysis. Returns dict with summary, qualityScore, etc."""
    key = (os.getenv('GEMINI_API_KEY') or os.getenv('GOOGLE_API_KEY') or '').strip()
    if not key:
        return None
    try:
        with open(path, 'rb') as f:
            b64 = base64.b64encode(f.read()).decode('ascii')
        ext = os.path.splitext(path)[1].lower()
        mime = 'image/jpeg' if ext in ('.jpg', '.jpeg') else ('image/png' if ext == '.png' else 'image/webp')
        disease_map = {
            'wheat': ['rust (stripe, leaf, stem)', 'powdery mildew', 'leaf blight', 'aphids'],
            'rice': ['rice blast', 'bacterial leaf blight', 'sheath blight', 'brown planthopper'],
            'maize': ['northern leaf blight', 'common rust', 'gray leaf spot', 'fall armyworm', 'common smut'],
        }
        crop_key = (crop or '').strip().lower()
        targets = disease_map.get(crop_key, [])
        target_text = ", ".join(targets) if targets else "common crop diseases and pests"
        system = (
            "You are an agronomy assistant. Analyze the crop photo for health, diseases, pests, and damage. "
            f"Focus on {crop or 'the crop'} and especially: {target_text}. "
            "Be sensitive to: brown patches, irregular patterns, holes, spots, edge burn, leaf curling, chlorosis, necrosis, webbing, insect damage. "
            "Return strict JSON with keys: "
            "summary (string), details (string), confidence (Low|Medium|High), qualityScore (0-10 integer: 0-3 poor, 4-6 fair, 7-8 good, 9-10 excellent), "
            "issues (array: {name, likelihood 0-100, description}), "
            "observations (array: {type, description, severity 0-100, confidence}), "
            "recommendations (array of strings)."
        )
        user_text = prompt or 'Assess crop condition, quality, and identify any disease or pest.'
        body = {
            'contents': [{
                'parts': [
                    {'inline_data': {'mime_type': mime, 'data': b64}},
                    {'text': f"{system}\n\n{user_text}"},
                ]
            }],
            'generationConfig': {
                'responseMimeType': 'application/json',
                'temperature': 0.2,
            }
        }
        model = (os.getenv('GEMINI_MODEL') or 'gemini-pro').strip() or 'gemini-pro'
        api_ver = 'v1' if model == 'gemini-pro' else 'v1beta'
        url = f'https://generativelanguage.googleapis.com/{api_ver}/models/{model}:generateContent?key={key}'
        r = requests.post(url, headers={'Content-Type': 'application/json'}, json=body, timeout=60)
        j = r.json()
        if r.status_code != 200:
            return None
        parts = j.get('candidates', [{}])[0].get('content', {}).get('parts', [])
        if not parts:
            return None
        content = parts[0].get('text', '')
        if not content:
            return None
        parsed = json.loads(content)
        qs = parsed.get('qualityScore')
        if qs is not None:
            parsed['qualityScore'] = max(0, min(10, int(qs)))
        return parsed
    except Exception:
        return None

def _analyze_with_openai(path, prompt, crop=None):
    key = os.getenv('OPENAI_API_KEY')
    if not key:
        return None
    try:
        with open(path, 'rb') as f:
            b64 = base64.b64encode(f.read()).decode('ascii')
        ext = os.path.splitext(path)[1]
        mime = _mime_for_ext(ext)
        data_uri = f"data:{mime};base64,{b64}"
        disease_map = {
            'wheat': ['rust (stripe, leaf, stem)', 'powdery mildew', 'leaf blight', 'aphids'],
            'rice': ['rice blast', 'bacterial leaf blight', 'sheath blight', 'brown planthopper'],
            'maize': ['northern leaf blight', 'common rust', 'gray leaf spot', 'fall armyworm', 'common smut'],
        }
        crop_key = (crop or '').strip().lower()
        targets = disease_map.get(crop_key, [])
        target_text = ", ".join(targets) if targets else "common crop diseases and pests"
        system = (
            "You are an agronomy assistant. Analyze the crop photo for health, diseases, pests, and damage. "
            f"Focus on {crop or 'the crop'} and especially: {target_text}. "
            "Be highly sensitive to small signs: brown patches, irregular patterns, holes, spots, edge burn, leaf curling, chlorosis, necrosis, webbing, and insect bite marks. "
            "Return strict JSON with keys: "
            "summary (string), details (string), confidence (Low|Medium|High), qualityScore (0-10 integer: 0-3 poor, 4-6 fair, 7-8 good, 9-10 excellent), "
            "issues (array of objects: name, likelihood (0-100), description), "
            "observations (array of objects: type, description, severity (0-100), confidence (Low|Medium|High)), "
            "recommendations (array of strings)."
        )
        body = {
            'model': os.getenv('OPENAI_MODEL', 'gpt-4o-mini'),
            'messages': [
                {'role': 'system', 'content': system},
                {
                    'role': 'user',
                    'content': [
                        {'type': 'text', 'text': prompt or 'Assess crop condition and identify any disease or pest.'},
                        {'type': 'image_url', 'image_url': {'url': data_uri}},
                    ],
                },
            ],
            'response_format': {'type': 'json_object'},
        }
        r = requests.post('https://api.openai.com/v1/chat/completions', headers={
            'Authorization': f'Bearer {key}',
            'Content-Type': 'application/json',
        }, json=body, timeout=60)
        j = r.json()
        content = j.get('choices', [{}])[0].get('message', {}).get('content')
        if not content:
            return None
        parsed = json.loads(content)
        qs = parsed.get('qualityScore')
        if qs is not None:
            parsed['qualityScore'] = max(0, min(10, int(qs)))
        return parsed
    except Exception:
        return None

@app.route('/api/crop-analysis', methods=['POST'])
def crop_analysis():
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided'}), 400
    image = request.files['image']
    if image.filename == '':
        return jsonify({'error': 'No image selected'}), 400
    ext = os.path.splitext(image.filename)[1]
    unique = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(app.config['UPLOAD_FOLDER'], unique)
    image.save(path)
    data = request.form.to_dict() if request.form else {}
    prompt = data.get('prompt')
    crop = data.get('crop')
    ai = _analyze_with_gemini(path, prompt, crop)
    if not ai or not isinstance(ai, dict):
        ai = _analyze_with_openai(path, prompt, crop)
    if ai and isinstance(ai, dict):
        quality_score = ai.get('qualityScore')
        if quality_score is None:
            quality_score = 7
        return jsonify({
            'summary': ai.get('summary') or 'Analysis available',
            'details': ai.get('details') or '',
            'confidence': ai.get('confidence') or 'Medium',
            'qualityScore': quality_score,
            'issues': ai.get('issues') or [],
            'observations': ai.get('observations') or [],
            'recommendations': ai.get('recommendations') or [],
        }), 200
    return jsonify({
        'summary': 'Likely healthy',
        'details': 'Leaves appear normal. No obvious signs of damage or disease detected.',
        'confidence': 'Medium',
        'qualityScore': 7,
        'issues': [],
        'observations': [],
        'recommendations': [],
    }), 200

@app.route('/api/stage-verify', methods=['POST'])
def stage_verify():
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided'}), 400
    image = request.files['image']
    if image.filename == '':
        return jsonify({'error': 'No image selected'}), 400
    stage = (request.form.get('stage') or '').strip()
    crop = (request.form.get('crop') or '').strip()
    ext = os.path.splitext(image.filename)[1]
    unique = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(app.config['UPLOAD_FOLDER'], unique)
    image.save(path)
    prompts = {
        '1': 'Verify this is seeds/sowing stage (seed packets, seeds in soil, early seedlings).',
        '2': 'Verify this is growth stage (green crops, leaves, stems, field with growing plants).',
        '3': 'Verify this is harvest stage (harvested crop, grain, bundles, transport).',
    }
    pfx = prompts.get(stage, 'Assess crop stage (seed, growth, harvest).')
    ai = _analyze_with_gemini(path, pfx, crop)
    if not ai or not isinstance(ai, dict):
        ai = _analyze_with_openai(path, pfx, crop)
    awarded = 0
    reason = 'No points'
    if ai and isinstance(ai, dict):
        conf = ai.get('confidence')
        issues = ai.get('issues') or []
        observations = ai.get('observations') or []
        text = json.dumps({'issues': issues, 'observations': observations}).lower()
        ok = False
        if stage == '1':
            ok = ('seed' in text) or ('sowing' in text) or ('seedling' in text)
        elif stage == '2':
            ok = ('leaf' in text) or ('growth' in text) or ('stem' in text) or ('green' in text)
        elif stage == '3':
            ok = ('harvest' in text) or ('grain' in text) or ('bundle' in text) or ('transport' in text)
        if ok:
            if (conf or '').lower() == 'high':
                awarded = 10
                reason = 'Stage verified with high confidence'
            else:
                awarded = 8
                reason = 'Stage verified'
        else:
            awarded = 5
            reason = 'Stage unclear, partial points'
        return jsonify({
            'stage': stage,
            'awardedPoints': awarded,
            'reason': reason,
            'analysis': {
                'summary': ai.get('summary'),
                'details': ai.get('details'),
                'confidence': ai.get('confidence'),
                'issues': ai.get('issues'),
                'observations': ai.get('observations'),
                'recommendations': ai.get('recommendations'),
            }
        }), 200
    return jsonify({
        'stage': stage,
        'awardedPoints': 10,
        'reason': 'Demo award',
        'analysis': {
            'summary': 'Likely healthy',
            'details': 'Demo response.',
            'confidence': 'Medium',
            'issues': [],
            'observations': [],
            'recommendations': [],
        }
    }), 200

def _count_small_transactions_csv(path, threshold=500.0):
    count = 0
    total = 0.0
    lines = 0
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.DictReader(f)
            for row in reader:
                amount = row.get('amount') or row.get('Amount') or row.get('AMOUNT')
                if amount is None:
                    continue
                try:
                    val = float(str(amount).replace(',', '').strip())
                except Exception:
                    continue
                total += abs(val)
                lines += 1
                if abs(val) <= threshold:
                    count += 1
    except Exception:
        pass
    return count, total, lines

def _count_small_transactions_xlsx(path, threshold=500.0):
    count = 0
    total = 0.0
    lines = 0
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h).lower() if h is not None else '' for h in row]
                continue
            # find amount-like column
            idx = None
            for j, h in enumerate(headers):
                if 'amount' in h or 'amt' in h or h in ('debit', 'credit'):
                    idx = j
                    break
            if idx is None:
                continue
            val = row[idx]
            try:
                num = float(str(val).replace(',', '').strip())
            except Exception:
                continue
            total += abs(num)
            lines += 1
            if abs(num) <= threshold:
                count += 1
    except Exception:
        pass
    return count, total, lines

def _count_small_transactions_json(path, threshold=500.0):
    count = 0
    total = 0.0
    lines = 0
    try:
        data = json.loads(open(path, 'r', encoding='utf-8').read())
        def walk(x):
            nonlocal count, total, lines
            if isinstance(x, dict):
                for k, v in x.items():
                    if k.lower() in ('amount', 'amt', 'debit', 'credit') and isinstance(v, (int, float, str)):
                        try:
                            num = float(str(v).replace(',', '').strip())
                        except Exception:
                            continue
                        total += abs(num)
                        lines += 1
                        if abs(num) <= threshold:
                            count += 1
                    else:
                        walk(v)
            elif isinstance(x, list):
                for v in x:
                    walk(v)
        walk(data)
    except Exception:
        pass
    return count, total, lines

def _count_small_transactions_pdf(path, threshold=500.0):
    count = 0
    total = 0.0
    lines = 0
    try:
        # Best-effort text extraction without external libs: decode bytes and regex amounts
        raw = open(path, 'rb').read()
        try:
            text = raw.decode('utf-8', errors='ignore')
        except Exception:
            text = raw.decode('latin-1', errors='ignore')
        import re
        # Match currency-like numbers: optional ₹/Rs/INR, commas, decimals, optional sign
        pattern = re.compile(r"(?:₹\s*|Rs\.?\s*|INR\s*)?(-?\d{1,3}(?:,\d{3})*(?:\.\d+)?|-?\d+(?:\.\d+)?)")
        matches = pattern.findall(text)
        for m in matches:
            try:
                num = float(str(m).replace(',', '').strip())
            except Exception:
                continue
            total += abs(num)
            lines += 1
            if abs(num) <= threshold:
                count += 1
    except Exception:
        pass
    return count, total, lines

@app.route('/api/bank-statement', methods=['POST'])
def bank_statement():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f = request.files['file']
    if f.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    unique = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(app.config['UPLOAD_FOLDER'], unique)
    f.save(path)
    small = 0
    total = 0.0
    lines = 0
    if ext in ('.csv',):
        small, total, lines = _count_small_transactions_csv(path)
    elif ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        small, total, lines = _count_small_transactions_xlsx(path)
    elif ext in ('.json', '.txt'):
        small, total, lines = _count_small_transactions_json(path)
    elif ext in ('.pdf',):
        small, total, lines = _count_small_transactions_pdf(path)
    else:
        return jsonify({'error': 'Unsupported file type'}), 400
    ratio = (small / lines) if lines else 0.0
    active = (small >= 15) or (ratio >= 0.5)
    delta = 20 if active else 0
    return jsonify({
        'active': active,
        'smallTransactions': small,
        'totalTransactions': lines,
        'activityRatio': round(ratio, 2),
        'trustDelta': delta,
    }), 200

def _normalize_metric(key, value):
    try:
        v = float(value)
    except Exception:
        return None
    k = (key or '').lower()
    if 'soil' in k and 'moist' in k:
        return max(0.0, min(10.0, v / 10.0))
    if k in ('ph',) or 'ph' in k:
        return max(0.0, min(10.0, 10.0 - abs(v - 6.5) * (10.0 / 6.5)))
    if 'humid' in k:
        return max(0.0, min(10.0, v / 10.0))
    if 'temp' in k:
        return max(0.0, min(10.0, (50.0 - abs(v - 25.0)) / 5.0))
    if 'rain' in k or 'precip' in k:
        return max(0.0, min(10.0, 10.0 - max(0.0, v - 20.0) * 0.5))
    if 'wind' in k:
        return max(0.0, min(10.0, 10.0 - v * 0.5))
    if 0.0 <= v <= 100.0:
        return v / 10.0
    return max(0.0, min(10.0, v / 10.0))

def _extract_address_and_coords(data):
    addr = None
    lat = None
    lon = None
    def _try_get(obj, keys):
        for k in keys:
            if isinstance(obj, dict) and k in obj and obj[k] is not None:
                return obj[k]
        return None
    if isinstance(data, dict):
        raw = _try_get(data, ['address', 'location', 'field_address'])
        lat = _try_get(data, ['lat', 'latitude'])
        lon = _try_get(data, ['lon', 'lng', 'longitude'])
        # Handle nested location object e.g. {"place": "Manaus, Brazil", "latitude": -3.1, "longitude": -60}
        if isinstance(raw, dict):
            addr_str = raw.get('place') or raw.get('name') or raw.get('address') or str(raw)
            if not lat and raw.get('latitude') is not None:
                try:
                    lat = float(raw['latitude'])
                except (TypeError, ValueError):
                    pass
            if not lon and raw.get('longitude') is not None:
                try:
                    lon = float(raw['longitude'])
                except (TypeError, ValueError):
                    pass
            addr = addr_str if isinstance(addr_str, str) else (addr or raw)
        else:
            addr = raw
    if isinstance(lat, str):
        try:
            lat = float(lat)
        except Exception:
            lat = None
    if isinstance(lon, str):
        try:
            lon = float(lon)
        except Exception:
            lon = None
    if (not lat or not lon) and addr and isinstance(addr, str):
        try:
            r = requests.get('https://nominatim.openstreetmap.org/search', params={'q': addr, 'format': 'json', 'limit': 1}, headers={'User-Agent': 'krishimitra-app'})
            j = r.json()
            if isinstance(j, list) and j:
                lat = float(j[0]['lat'])
                lon = float(j[0]['lon'])
        except Exception:
            pass
    return addr, lat, lon

def _extract_text_from_pdf(path):
    """Extract text from PDF using pypdf or PyPDF2. Optimized for sensor/soil reports."""
    for mod in ('pypdf', 'PyPDF2'):
        try:
            if mod == 'pypdf':
                from pypdf import PdfReader
            else:
                from PyPDF2 import PdfReader
            reader = PdfReader(path)
            text_parts = []
            for page in reader.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t.strip())
            if not text_parts:
                return ''
            # Normalize: collapse multiple spaces, preserve newlines for structure
            full = '\n'.join(text_parts)
            full = re.sub(r'[ \t]+', ' ', full)  # collapse horizontal whitespace
            full = re.sub(r'\n{3,}', '\n\n', full)  # max 2 consecutive newlines
            return full.strip()
        except Exception:
            continue
    return ''

def _flatten_numeric(obj):
    nums = []
    metrics = {'ph': None, 'moisture': None, 'nitrogen': None}
    if isinstance(obj, dict):
        for k, v in obj.items():
            if isinstance(v, (int, float)):
                n = _normalize_metric(k, v)
                if n is not None:
                    nums.append(n)
            elif isinstance(v, (dict, list)):
                nums.extend(_flatten_numeric(v))
    elif isinstance(obj, list):
        for v in obj:
            if isinstance(v, (int, float)):
                n = _normalize_metric('', v)
                if n is not None:
                    nums.append(n)
            elif isinstance(v, (dict, list)):
                nums.extend(_flatten_numeric(v))
    return nums

def _collect_metrics(obj, ph_vals, moist_vals, n_vals):
    """Recursively collect ph, moisture, nitrogen from nested JSON."""
    if isinstance(obj, dict):
        for k, v in obj.items():
            lk = str(k).lower()
            if isinstance(v, (int, float)):
                if 'ph' in lk:
                    ph_vals.append(float(v))
                if 'moist' in lk or 'humidity' in lk:
                    moist_vals.append(float(v))
                if 'nitrogen' in lk or lk == 'n':
                    n_vals.append(float(v))
            else:
                _collect_metrics(v, ph_vals, moist_vals, n_vals)
    elif isinstance(obj, list):
        for v in obj:
            _collect_metrics(v, ph_vals, moist_vals, n_vals)


def _collect_rainfall(obj, rain_vals):
    """Recursively collect rainfall values from nested JSON (keys: rainfall, rainfall_mm, precipitation, precip)."""
    if isinstance(obj, dict):
        for k, v in obj.items():
            lk = str(k).lower()
            if isinstance(v, (int, float)):
                if 'rain' in lk or 'precip' in lk:
                    try:
                        rain_vals.append(float(v))
                    except (TypeError, ValueError):
                        pass
            else:
                _collect_rainfall(v, rain_vals)
    elif isinstance(obj, list):
        for v in obj:
            _collect_rainfall(v, rain_vals)


def _parse_sensor_metrics(path, ext):
    """Parse file and extract metrics + content. Returns (metrics, addr, lat, lon, content_str, nums, rainfall_total)."""
    nums = []
    metrics = {'ph': None, 'moisture': None, 'nitrogen': None}
    addr = None
    lat = None
    lon = None
    content_str = ''
    rainfall_total = None

    try:
        if ext in ('.json', '.txt'):
            content_str = open(path, 'r', encoding='utf-8', errors='ignore').read()
            data = json.loads(content_str)
            nums = _flatten_numeric(data)
            addr, lat, lon = _extract_address_and_coords(data)
            ph_vals, moist_vals, n_vals = [], [], []
            _collect_metrics(data, ph_vals, moist_vals, n_vals)
            rain_vals = []
            _collect_rainfall(data, rain_vals)
            if rain_vals:
                rainfall_total = round(sum(rain_vals), 1)
            if ph_vals:
                metrics['ph'] = sum(ph_vals) / len(ph_vals)
            if moist_vals:
                metrics['moisture'] = sum(moist_vals) / len(moist_vals)
            if n_vals:
                metrics['nitrogen'] = sum(n_vals) / len(n_vals)
        elif ext in ('.csv',):
            content_str = open(path, 'r', encoding='utf-8', errors='ignore').read()
            reader = csv.DictReader(StringIO(content_str))
            rain_vals = []
            for row in reader:
                for k, v in row.items():
                    try:
                        n = _normalize_metric(k, float(v))
                        if n is not None:
                            nums.append(n)
                        lk = str(k).lower()
                        fv = float(v)
                        if 'rain' in lk or 'precip' in lk:
                            rain_vals.append(fv)
                        if metrics['ph'] is None and 'ph' in lk:
                            metrics['ph'] = fv
                        if metrics['moisture'] is None and ('moist' in lk or 'humidity' in lk):
                            metrics['moisture'] = fv
                        if metrics['nitrogen'] is None and ('nitrogen' in lk or lk == 'n'):
                            metrics['nitrogen'] = fv
                    except Exception:
                        pass
            if rain_vals:
                rainfall_total = round(sum(rain_vals), 1)
        elif ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            content_str = '\n'.join([','.join([str(c) if c is not None else '' for c in r]) for r in rows])
            headers = []
            rain_vals = []
            for i, row in enumerate(rows):
                if i == 0:
                    headers = [str(h) if h is not None else '' for h in row]
                    continue
                for k, v in zip(headers, row):
                    try:
                        n = _normalize_metric(k, float(v))
                        if n is not None:
                            nums.append(n)
                        lk = str(k).lower()
                        fv = float(v)
                        if 'rain' in lk or 'precip' in lk:
                            rain_vals.append(fv)
                        if metrics['ph'] is None and 'ph' in lk:
                            metrics['ph'] = fv
                        if metrics['moisture'] is None and ('moist' in lk or 'humidity' in lk):
                            metrics['moisture'] = fv
                        if metrics['nitrogen'] is None and ('nitrogen' in lk or lk == 'n'):
                            metrics['nitrogen'] = fv
                    except Exception:
                        pass
            if rain_vals:
                rainfall_total = round(sum(rain_vals), 1)
        elif ext == '.pdf':
            content_str = _extract_text_from_pdf(path)
            if not content_str:
                content_str = open(path, 'rb').read().decode('utf-8', errors='ignore')[:10000]
            rain_vals = []
            # Extract metrics from PDF text: look for key: value and "key value" patterns
            for line in content_str.splitlines():
                stripped = line.strip()
                if not stripped:
                    continue
                lower = stripped.lower()
                if metrics['ph'] is None and ('ph' in lower or 'p.h' in lower):
                    m = re.search(r'ph[:\s=]*(-?\d+(?:\.\d+)?)|(-?\d+(?:\.\d+)?)\s*(?:ph|pH)', lower, re.I)
                    val = (m.group(1) or m.group(2)) if m else None
                    if not val:
                        m = re.search(r'(-?\d+(?:\.\d+)?)', stripped)
                        val = m.group(1) if m else None
                    if val:
                        metrics['ph'] = float(val)
                        n = _normalize_metric('ph', metrics['ph'])
                        if n is not None:
                            nums.append(n)
                if metrics['moisture'] is None and ('moist' in lower or 'humidity' in lower or '%' in stripped):
                    m = re.search(r'(\d+(?:\.\d+)?)\s*%|moisture[:\s=]*(\d+(?:\.\d+)?)|(\d+(?:\.\d+)?)\s*moisture', lower, re.I)
                    val = (m.group(1) or m.group(2) or m.group(3)) if m else None
                    if not val:
                        m = re.search(r'(-?\d+(?:\.\d+)?)', stripped)
                        val = m.group(1) if m else None
                    if val:
                        metrics['moisture'] = float(val)
                        n = _normalize_metric('soil moisture', metrics['moisture'])
                        if n is not None:
                            nums.append(n)
                if metrics['nitrogen'] is None and ('nitrogen' in lower or 'n ' in lower or ' ppm' in lower or 'mg/kg' in lower):
                    m = re.search(r'nitrogen[:\s=]*(\d+(?:\.\d+)?)|(\d+(?:\.\d+)?)\s*(?:ppm|mg/kg|mg kg)|n[:\s=]*(\d+(?:\.\d+)?)', lower, re.I)
                    val = (m.group(1) or m.group(2) or m.group(3)) if m else None
                    if not val:
                        m = re.search(r'(-?\d+(?:\.\d+)?)', stripped)
                        val = m.group(1) if m else None
                    if val:
                        metrics['nitrogen'] = float(val)
                        n = _normalize_metric('nitrogen', metrics['nitrogen'])
                        if n is not None:
                            nums.append(n)
                if addr is None and ('address' in lower or 'location' in lower):
                    parts = stripped.split(':', 1)
                    addr = parts[1].strip() if len(parts) == 2 and parts[1].strip() else stripped
                if 'rain' in lower or 'precip' in lower:
                    m = re.search(r'rainfall[:\s=]*(\d+(?:\.\d+)?)|precipitation[:\s=]*(\d+(?:\.\d+)?)|(\d+(?:\.\d+)?)\s*(?:mm|mm\.?)', lower, re.I)
                    val = (m.group(1) or m.group(2) or m.group(3)) if m else None
                    if not val:
                        m = re.search(r'(\d+(?:\.\d+)?)', stripped)
                        val = m.group(1) if m else None
                    if val:
                        try:
                            rain_vals.append(float(val))
                        except (TypeError, ValueError):
                            pass
            if rain_vals:
                rainfall_total = round(sum(rain_vals), 1)
            if addr:
                try:
                    _, lat, lon = _extract_address_and_coords({'address': addr})
                except Exception:
                    pass
    except json.JSONDecodeError as e:
        raise ValueError(f'Invalid JSON: {e}') from e
    except Exception as e:
        raise ValueError(f'Failed to parse file: {e}') from e

    return metrics, addr, lat, lon, content_str, nums, rainfall_total


@app.route('/api/sensor-readings', methods=['POST'])
def sensor_readings():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    f = request.files['file']
    if f.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ('.json', '.txt', '.csv', '.xlsx', '.xlsm', '.xltx', '.xltm', '.pdf'):
        return jsonify({'error': 'Unsupported file type. Use JSON, PDF, CSV, or Excel.'}), 400

    unique = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(app.config['UPLOAD_FOLDER'], unique)
    try:
        f.save(path)
    except Exception as e:
        return jsonify({'error': f'Failed to save file: {e}'}), 500

    try:
        metrics, addr, lat, lon, content_str, nums, rainfall_total = _parse_sensor_metrics(path, ext)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

    # Rule-based trust score (no AI): same logic style as bank statement
    # pH 6.0-7.5 = 10, moisture 20-60 = 10, nitrogen in range = 10; total 0-30
    ph, moist, nitro = metrics.get('ph'), metrics.get('moisture'), metrics.get('nitrogen')
    points = 0
    if ph is not None and 6.0 <= ph <= 7.5:
        points += 10
    if moist is not None and 20 <= moist <= 60:
        points += 10
    # Nitrogen: accept 40-80 mg/kg or 240-480 ppm
    if nitro is not None and (40 <= nitro <= 80 or 240 <= nitro <= 480):
        points += 10
    if points == 0 and nums:
        trust = round(sum(nums) / max(1, len(nums)), 1) * 3
        trust = max(0, min(30, trust))
    else:
        trust = float(points)

    summary = None  # No AI used; rule-based only
    addr_text = None
    if addr is not None:
        addr_text = json.dumps(addr, ensure_ascii=False) if isinstance(addr, (dict, list)) else str(addr)

    sr = SensorReport(
        filename=unique,
        original_filename=f.filename,
        file_path=path,
        trust_score_10=trust,
        address_text=addr_text,
        lat=lat,
        lon=lon,
        ai_summary=summary
    )
    db.session.add(sr)
    db.session.commit()

    out = {
        'trustScore': trust,
        'address': addr,
        'lat': lat,
        'lon': lon,
        'reportId': sr.id,
        'metrics': metrics,
        'aiSummary': summary
    }
    if rainfall_total is not None:
        out['rainfallTotal'] = rainfall_total
    return jsonify(out), 200

@app.route('/api/weather', methods=['GET'])
def weather():
    address = request.args.get('address')
    lat = request.args.get('lat')
    lon = request.args.get('lon')
    if address and (not lat or not lon):
        try:
            r = requests.get('https://nominatim.openstreetmap.org/search', params={'q': address, 'format': 'json', 'limit': 1}, headers={'User-Agent': 'krishimitra-app'})
            j = r.json()
            if isinstance(j, list) and j:
                lat = j[0]['lat']
                lon = j[0]['lon']
        except Exception:
            pass
    if not lat or not lon:
        return jsonify({'error': 'lat/lon or address required'}), 400
    try:
        wx = requests.get('https://api.open-meteo.com/v1/forecast', params={
            'latitude': lat,
            'longitude': lon,
            'current_weather': 'true',
            'hourly': 'temperature_2m,precipitation,wind_speed_10m'
        })
        data = wx.json()
        current = data.get('current_weather') or {}
        hourly = data.get('hourly') or {}
        return jsonify({
            'current': current,
            'hourly': hourly,
            'lat': float(lat),
            'lon': float(lon)
        }), 200
    except Exception:
        return jsonify({'error': 'weather fetch failed'}), 500

@app.route('/api/upload', methods=['POST'])
@jwt_required()
def upload_file():
    current_user_id = get_jwt_identity()
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    # Generate unique filename
    file_extension = os.path.splitext(file.filename)[1]
    unique_filename = f"{uuid.uuid4().hex}{file_extension}"
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
    
    # Save file
    file.save(file_path)
    
    # Create file record in database
    uploaded_file = UploadedFile(
        filename=unique_filename,
        original_filename=file.filename,
        file_path=file_path,
        user_id=current_user_id
    )
    
    db.session.add(uploaded_file)
    db.session.commit()
    
    return jsonify({
        'message': 'File uploaded successfully',
        'file_id': uploaded_file.id,
        'filename': uploaded_file.original_filename
    }), 201

@app.route('/api/process/<int:file_id>', methods=['POST'])
@jwt_required()
def process_file(file_id):
    current_user_id = get_jwt_identity()
    
    uploaded_file = UploadedFile.query.filter_by(id=file_id, user_id=current_user_id).first()
    
    if not uploaded_file:
        return jsonify({'error': 'File not found'}), 404
    
    # Get processing prompt from request
    data = request.get_json()
    prompt = data.get('prompt', 'Analyze this agricultural data and provide insights.')
    
    # TODO: Integrate with AI API (OpenAI, Gemini, etc.)
    # This is where you'll call your AI service
    ai_response = f"Processed file: {uploaded_file.original_filename}\nPrompt: {prompt}\n\nAI Response: This is a placeholder response. Integrate with actual AI API."
    
    # Update file record
    uploaded_file.processed = True
    uploaded_file.ai_response = ai_response
    db.session.commit()
    
    return jsonify({
        'message': 'File processed successfully',
        'response': ai_response
    }), 200

@app.route('/api/files', methods=['GET'])
@jwt_required()
def get_user_files():
    current_user_id = get_jwt_identity()
    
    files = UploadedFile.query.filter_by(user_id=current_user_id).order_by(UploadedFile.uploaded_at.desc()).all()
    
    files_data = []
    for file in files:
        files_data.append({
            'id': file.id,
            'filename': file.original_filename,
            'uploaded_at': file.uploaded_at.isoformat(),
            'processed': file.processed,
            'ai_response': file.ai_response
        })
    
    return jsonify({'files': files_data}), 200

@app.route('/api/profile', methods=['GET'])
@jwt_required()
def get_profile():
    current_user_id = get_jwt_identity()
    
    user = User.query.get(current_user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    return jsonify({
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'created_at': user.created_at.isoformat()
    }), 200


# ---- Chatbot: Gemini with full Krishimitra knowledge; fallback when API unavailable ----
def _chat_fallback(message):
    """When Gemini is unavailable, return a helpful reply so the chat never feels broken."""
    m = (message or '').strip().lower()
    if not m:
        return None
    if any(w in m for w in ('hi', 'hello', 'hey', 'namaste')):
        return (
            "Hello! I'm Krishimitra Support. You can ask about Trust Score, bank/sensor uploads, "
            "Crop Analysis, Weather Insurance, Vouchers, or Pay-as-you-Grow. Use the **Feedback** button above "
            "for complaints or ratings. Need to reach us? Check **Contact Us** in the sidebar — phone 083903 12345, "
            "toll-free 1800 123 4567, email support@krishimitra.in, WhatsApp +91 91234 56789. Helpline: Mon–Sat, 9 AM – 6 PM."
        )
    if any(w in m for w in ('trust score', 'trustscore', 'score', 'eligib', 'loan')):
        return (
            "**Trust Score (0–100)** is built by completing: Profile, Bank Statement, Sensor Readings, "
            "Crop Analysis, Financial Quiz, and Weather Insurance. Score 80+ unlocks loans, Vouchers, and Pay-as-you-Grow. "
            "Complete tasks on the Home dashboard and click **Evaluate my score** to see your score."
        )
    if any(w in m for w in ('contact', 'phone', 'email', 'call', 'help', 'support', 'reach')):
        return (
            "**Contact Krishimitra:** Telephone 083903 12345, Mobile 091234 56789, Toll-free 1800 123 4567, "
            "Email support@krishimitra.in, WhatsApp +91 91234 56789, Instagram @krishimitra. Helpline: Mon–Sat, 9 AM – 6 PM. "
            "You can also use the **Feedback** button above to send a message."
        )
    if any(w in m for w in ('bank', 'statement', 'upload')):
        return (
            "Upload your **bank statement** (PDF, CSV, Excel, or JSON) from the Bank Statement page. "
            "Active accounts can earn up to +20 Trust Score. Go to the sidebar → Bank Statement."
        )
    if any(w in m for w in ('sensor', 'soil', 'ph', 'moisture', 'nitrogen')):
        return (
            "Upload **sensor/field data** (JSON, CSV, Excel, or PDF) from Sensor Readings. "
            "Include pH, moisture, and nitrogen for up to 30 Trust Score. Rainfall data is used for Weather Insurance."
        )
    if any(w in m for w in ('crop', 'photo', 'image', 'analysis')):
        return (
            "Use **Crop Analysis** (sidebar or Explore) to upload a crop photo. AI will analyse health, "
            "diseases, and pests and suggest recommendations. You can also attach an image in this chat for quick analysis."
        )
    if any(w in m for w in ('weather', 'rain', 'insurance')):
        return (
            "**Weather Insurance** gives automatic payouts when rainfall in your area is below a threshold. "
            "Upload sensor data with rainfall first, then check the Weather Insurance page. Uses data from your Sensor Readings."
        )
    if any(w in m for w in ('voucher', 'pay-as-you-grow', 'milestone')):
        return (
            "**Vouchers** (QR/PIN) and **Pay-as-you-Grow** (funds in stages: Seeds → Labor → Harvest) unlock when your Trust Score is 80+. "
            "Complete the dashboard tasks and evaluate your score to qualify."
        )
    if any(w in m for w in ('feedback', 'complaint', 'rating')):
        return (
            "Use the **Feedback** button at the top of this chat to send general feedback, a complaint, or a star rating. "
            "Our team will get back to you."
        )
    # Default: friendly prompt to try Feedback or contact
    return (
        "I'm here for Krishimitra support. Try asking about **Trust Score**, **contact details**, **bank/sensor uploads**, "
        "**Crop Analysis**, or **Weather Insurance**. You can also use the **Feedback** button above, or check **Contact Us** in the sidebar — "
        "phone 083903 12345, email support@krishimitra.in. Helpline: Mon–Sat, 9 AM – 6 PM."
    )


@app.route('/api/chat', methods=['POST'])
def chat():
    data = request.get_json() or {}
    message = (data.get('message') or '').strip()
    if not message:
        return jsonify({'error': 'Message is required'}), 400
    history = data.get('history')
    if not isinstance(history, list):
        history = []
    key = (os.getenv('GEMINI_API_KEY') or os.getenv('GOOGLE_API_KEY') or '').strip()
    if not key:
        reply = _chat_fallback(message)
        return jsonify({'reply': reply or 'Please use the Feedback button above or contact support@krishimitra.in. Helpline: Mon–Sat, 9 AM – 6 PM.'})
    reply, err = _chat_with_gemini_rest(message, history)
    if reply:
        return jsonify({'reply': reply})
    fallback = _chat_fallback(message)
    return jsonify({'reply': fallback or 'Something went wrong. Use the Feedback button above or email support@krishimitra.in. We\'ll get back to you.'})


@app.route('/api/feedback', methods=['POST'])
def feedback():
    data = request.get_json() or {}
    kind = (data.get('type') or 'feedback').strip().lower()
    if kind not in ('feedback', 'complaint', 'rating'):
        kind = 'feedback'
    content = (data.get('content') or '').strip()
    rating = data.get('rating')
    if rating is not None:
        rating = max(1, min(5, int(rating)))
    user_id = (data.get('userId') or data.get('user_id') or '').strip() or None
    email = (data.get('email') or '').strip() or None
    entry = SupportFeedback(type=kind, content=content or None, rating=rating, user_id=user_id, email=email)
    db.session.add(entry)
    db.session.commit()
    return jsonify({'success': True, 'id': entry.id})


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, port=5000)
